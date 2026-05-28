from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from dataclasses import dataclass, field
from typing import Optional


@dataclass
class ColumnDef:
    name: str
    width: Optional[float] = None
    font: Optional[dict] = None
    alignment: Optional[dict] = None


@dataclass
class TemplateFormat:
    """Structured description of a template .xlsx layout."""
    header_row: int = 4
    columns: list[ColumnDef] = field(default_factory=list)
    col_count: int = 0
    title_merged_range: Optional[str] = None
    title_font: Optional[dict] = None
    row_height: float = 100
    default_font: Optional[dict] = None


def read_template(path: str) -> TemplateFormat:
    wb = load_workbook(path)
    ws = wb.active

    fmt = TemplateFormat()

    # Find header row: scan row 1-10 for a row with "图片" or "品名" etc.
    header_row = find_header_row(ws)
    fmt.header_row = header_row

    # Title merge range (typically A1:? above header)
    for mr in ws.merged_cells.ranges:
        if mr.min_row < header_row and mr.max_col > 3:
            fmt.title_merged_range = str(mr)
            cell = ws.cell(mr.min_row, mr.min_col)
            fmt.title_font = {
                "name": cell.font.name,
                "size": cell.font.size,
                "bold": cell.font.bold,
            }
            break

    # Columns from header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col)
        name = (cell.value or "").strip()
        if not name and col > 13:
            continue
        col_letter = cell.column_letter
        col_def = ColumnDef(name=name)
        if col_letter in ws.column_dimensions:
            col_def.width = ws.column_dimensions[col_letter].width
        if cell.font:
            col_def.font = {
                "name": cell.font.name,
                "size": cell.font.size,
                "bold": cell.font.bold,
            }
        if cell.alignment:
            col_def.alignment = {
                "horizontal": cell.alignment.horizontal,
                "vertical": cell.alignment.vertical,
                "wrap_text": cell.alignment.wrap_text,
            }
        fmt.columns.append(col_def)

    fmt.col_count = len(fmt.columns)

    # Row height from first data row (header_row + 1)
    data_row = header_row + 1
    if data_row in ws.row_dimensions:
        fmt.row_height = ws.row_dimensions[data_row].height or 100

    # Default font from header cell
    hdr_cell = ws.cell(header_row, 1)
    fmt.default_font = {
        "name": hdr_cell.font.name or "宋体",
        "size": hdr_cell.font.size or 11,
        "bold": hdr_cell.font.bold or False,
    }

    wb.close()
    return fmt


def find_header_row(ws) -> int:
    """Scan for the row containing column headers. Look for common header keywords."""
    keywords = ["图片", "品名", "货号", "编码"]
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row, col).value or "")
            for kw in keywords:
                if kw in val:
                    return row
    return 4


def get_default_format() -> TemplateFormat:
    """Default format when no template is provided."""
    headers = ["图片", "品名", "编码", "备注", "件数", "每件数量", "总数量",
               "单价", "总金额", "货款", "运费", "重量/KG", "总重量/KG"]
    widths = [14.44, 19.11, 8, 20.55, 8, 10, 10, 8, 10, 10, 10, 9.78, 10]
    cols = []
    for h, w in zip(headers, widths):
        cols.append(ColumnDef(name=h, width=w,
                              font={"name": "宋体", "size": 11, "bold": True},
                              alignment={"horizontal": "center", "vertical": "center", "wrap_text": True}))
    return TemplateFormat(
        columns=cols,
        col_count=len(cols),
        title_merged_range="A1:M3",
        title_font={"name": "宋体", "size": 14, "bold": True},
        row_height=110,
        default_font={"name": "宋体", "size": 11, "bold": False},
    )
