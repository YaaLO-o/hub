import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from template_reader import TemplateFormat


def generate(template: TemplateFormat, data_rows: list[dict], photo_dir: str,
             output_path: str, title_text: str = ""):
    """
    Generate an order xlsx from template format + user data.

    data_rows: list of dicts with keys: filename, 品名, 单价, 件数
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    hr = template.header_row

    # Default title
    if not title_text:
        today = date.today()
        title_text = f"订货单         日期：{today.year}年{today.month}月{today.day}日"

    # Write title in merged range above header
    if template.title_merged_range:
        try:
            parts = template.title_merged_range.split(":")
            ws.merge_cells(template.title_merged_range)
            c = ws[parts[0]]
            c.value = title_text
            tf = template.title_font or {}
            c.font = Font(name=tf.get("name", "宋体"), size=tf.get("size", 14), bold=tf.get("bold", True))
            c.alignment = Alignment(horizontal="center", vertical="center")
        except Exception:
            pass

    # Write header row
    for ci, col_def in enumerate(template.columns, 1):
        c = ws.cell(row=hr, column=ci, value=col_def.name)
        f = col_def.font or {}
        c.font = Font(name=f.get("name", "宋体"), size=f.get("size", 11), bold=f.get("bold", True))
        a = col_def.alignment or {}
        c.alignment = Alignment(
            horizontal=a.get("horizontal", "center"),
            vertical=a.get("vertical", "center"),
            wrap_text=a.get("wrap_text", True),
        )
        c.border = thin

    # Set column widths
    for ci, col_def in enumerate(template.columns, 1):
        col_letter = ws.cell(row=hr, column=ci).column_letter
        if col_def.width:
            ws.column_dimensions[col_letter].width = col_def.width

    # Identify column indices by header name (flexible matching)
    field_map = {
        "图片": ["图片", "照片"],
        "品名": ["品名", "货号", "产品", "名称"],
        "编码": ["编码", "编号"],
        "备注": ["备注", "说明"],
        "件数": ["件数", "件"],
        "每件数量": ["每件数量", "每箱数量", "每箱"],
        "总数量": ["总数量", "总数", "数量合计", "总箱数"],
        "单价": ["单价"],
        "总金额": ["总金额", "金额", "合计金额"],
        "货款": ["货款"],
        "运费": ["运费"],
        "重量": ["重量", "毛重", "净重", "重量/KG"],
        "总重量": ["总重量", "总重", "总重量/KG"],
    }

    def col_idx(field):
        for kw in field_map.get(field, [field]):
            for i, col in enumerate(template.columns):
                if kw in col.name:
                    return i + 1
        return None

    col_image = col_idx("图片")
    col_name = col_idx("品名")
    col_code = col_idx("编码")
    col_note = col_idx("备注")
    col_pkg = col_idx("件数")
    col_qty_per = col_idx("每件数量")
    col_total_qty = col_idx("总数量")
    col_price = col_idx("单价")
    col_total_amt = col_idx("总金额")
    col_weight = col_idx("重量") or col_idx("总重量")
    col_total_weight = col_idx("总重量")
    col_payment = col_idx("货款")
    col_freight = col_idx("运费")
    col_count = template.col_count

    df = template.default_font or {}
    data_font = Font(name=df.get("name", "宋体"), size=df.get("size", 11), bold=False)

    # Write data rows
    for ri, row_data in enumerate(data_rows):
        row = hr + 1 + ri
        ws.row_dimensions[row].height = template.row_height or 110

        # For each column, fill data or empty
        for ci in range(1, col_count + 1):
            c = ws.cell(row=row, column=ci)
            c.font = data_font
            c.alignment = center
            c.border = thin

        # Known fields
        if col_name and row_data.get("品名"):
            ws.cell(row=row, column=col_name, value=row_data["品名"]).font = data_font
        if col_code and row_data.get("编码"):
            ws.cell(row=row, column=col_code, value=row_data["编码"]).font = data_font
        if col_note and row_data.get("备注"):
            ws.cell(row=row, column=col_note, value=row_data["备注"]).font = data_font
        if col_pkg:
            val = row_data.get("件数", 1)
            ws.cell(row=row, column=col_pkg, value=val).font = data_font
        if col_qty_per:
            val = row_data.get("每件数量", 240)
            ws.cell(row=row, column=col_qty_per, value=val).font = data_font
        if col_price:
            ws.cell(row=row, column=col_price, value=row_data.get("单价", 0)).font = data_font

        # Extra numeric fields from parsed data
        for col_var, field in [(col_weight, "重量"), (col_total_weight, "总重量"),
                                (col_payment, "货款"), (col_freight, "运费")]:
            if col_var and row_data.get(field) is not None:
                ws.cell(row=row, column=col_var, value=row_data[field]).font = data_font

        # Formulas
        if col_total_qty and col_pkg and col_qty_per:
            e = ws.cell(row=row, column=col_pkg)
            f = ws.cell(row=row, column=col_qty_per)
            ws.cell(row=row, column=col_total_qty, value=f"={e.coordinate}*{f.coordinate}")
        if col_total_amt and col_total_qty and col_price:
            g = ws.cell(row=row, column=col_total_qty)
            h = ws.cell(row=row, column=col_price)
            ws.cell(row=row, column=col_total_amt, value=f"={g.coordinate}*{h.coordinate}")

        # Insert image
        if col_image and row_data.get("filename") and photo_dir:
            img_path = os.path.join(photo_dir, row_data["filename"])
            if os.path.exists(img_path):
                try:
                    pil = PILImage.open(img_path)
                    cell_w = 100
                    ratio = cell_w / pil.width
                    new_w, new_h = cell_w, int(pil.height * ratio)
                    pil = pil.resize((new_w, new_h), PILImage.LANCZOS)

                    temp = os.path.join(os.environ['TEMP'], f"_order_img_{ri}.png")
                    pil.save(temp)
                    img = XLImage(temp)
                    img.width, img.height = new_w, new_h
                    img.anchor = ws.cell(row=row, column=col_image).coordinate
                    ws.add_image(img)
                except Exception:
                    pass

    # Total row — sums all numeric columns found in parsed data
    total_row = hr + 1 + len(data_rows)
    for ci in range(1, col_count + 1):
        ws.cell(row=total_row, column=ci).border = thin
        ws.cell(row=total_row, column=ci).font = Font(name=df.get("name", "宋体"), size=11, bold=True)
        ws.cell(row=total_row, column=ci).alignment = center

    # Collect which columns have numeric data worth summing
    sum_targets = []
    for field_name, col_idx_val in [
        ("总数量", col_total_qty), ("总金额", col_total_amt),
        ("总重量", col_total_weight), ("货款", col_payment), ("运费", col_freight),
    ]:
        if col_idx_val:
            # 总数量 and 总金额 are always summed (formulas exist);
            # other fields only if they have data
            if field_name in ("总数量", "总金额") or any(row_data.get(field_name) for row_data in data_rows):
                sum_targets.append((field_name, col_idx_val))

    if sum_targets:
        # Label in the leftmost column or before first sum column
        label_col = max(1, sum_targets[0][1] - 2) if sum_targets[0][1] > 2 else 1
        ws.cell(row=total_row, column=label_col, value="合计")
        ws.cell(row=total_row, column=label_col).font = Font(name=df.get("name", "宋体"), size=11, bold=True)

        first_data = hr + 1
        last_data = total_row - 1

        # SUM formulas for each numeric target
        for field_name, col_idx_val in sum_targets:
            ws.cell(row=total_row, column=col_idx_val,
                    value=f"=SUM({ws.cell(first_data, col_idx_val).coordinate}:{ws.cell(last_data, col_idx_val).coordinate})")

    wb.save(output_path)
    # Cleanup temp images
    for ri in range(len(data_rows)):
        p = os.path.join(os.environ['TEMP'], f"_order_img_{ri}.png")
        if os.path.exists(p):
            os.remove(p)
